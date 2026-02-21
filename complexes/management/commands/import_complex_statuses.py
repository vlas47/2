from argparse import ArgumentParser
from collections import Counter
from pathlib import Path

from django.core.management import BaseCommand, CommandError
from openpyxl import load_workbook

from complexes.models import PartnerComplex
from complexes.utils import normalize_complex_name, slugify_name


def _extract_fill_color(cell) -> str | None:
    if not cell:
        return None
    fg = getattr(cell.fill, 'fgColor', None)
    start = getattr(cell.fill, 'start_color', None)
    for color_source in (fg, start):
        if not color_source:
            continue
        rgb = getattr(color_source, 'rgb', None)
        if rgb:
            return rgb.upper()
    return None


class Command(BaseCommand):
    help = "Импортирует актуальность ЖК из Excel и сохраняет данные в PartnerComplex."

    def add_arguments(self, parser: ArgumentParser):
        parser.add_argument(
            "--source",
            required=True,
            help="Путь к Excel-файлу со списком ЖК (первая колонка с цветной заливкой).",
        )

    def handle(self, *args, **options):
        source = Path(options["source"])
        if not source.exists():
            raise CommandError(f"Файл {source} не найден.")
        workbook = load_workbook(source, read_only=True, data_only=True)
        sheet = workbook.active
        summary = Counter()
        updated = 0
        for row in sheet.iter_rows(min_row=2, max_col=1):
            cell = row[0]
            name = cell.value
            if not name:
                continue
            raw_name = str(name).strip()
            if not raw_name:
                continue
            normalized = normalize_complex_name(raw_name)
            color = _extract_fill_color(cell)
            state = "active"
            if color and color.endswith("FF0000"):
                state = "inactive"
            elif color and color.endswith("92D050"):
                state = "active"
            elif color and color != "00000000":
                state = "active"
            summary[state] += 1
            slug = slugify_name(raw_name)
            defaults = {
                "name": raw_name,
                "slug": slug,
                "is_active": state == "active",
                "normalized_name": normalized,
            }
            obj, created = PartnerComplex.objects.update_or_create(
                normalized_name=normalized,
                defaults=defaults,
            )
            if created or obj.is_active != defaults["is_active"] or obj.name != defaults["name"] or obj.slug != defaults["slug"]:
                obj.name = defaults["name"]
                obj.slug = defaults["slug"]
                obj.is_active = defaults["is_active"]
                obj.normalized_name = defaults["normalized_name"]
                obj.save()
            updated += 1
        self.stdout.write(f"Актуальных ЖК: {summary['active']}, Неактуальных: {summary['inactive']}")
